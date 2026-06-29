import os
import time
import pandas as pd # <--- Importiamo Pandas per la magia dell'esportazione Excel
from sklearn.svm import SVC
from sklearn.metrics import f1_score
from sklearn.model_selection import ParameterGrid
from tqdm import tqdm

def addestramento(nome_encoding, set_preparato, numero_features, modulo_encoding, iperparametri_quantistici, iperparametri_classici, nome_cartella):

    # ESTRAZIONE DEI DATI.

    train_set_preparato, val_set_preparato, _ = set_preparato

    _, y_train_preparato = train_set_preparato
    _, y_val_preparato = val_set_preparato

    # CREAZIONE DELLE GRIGLIE DI IPERPARAMETRI (quantistici e classici). 

    # Creiamo due griglie in modo da poter valutare i parametri quantistici, che richiedono il calcolo del kernel, da quelli classici, 
    # riducendo al minimo le volte in cui il kernel quantistico viene calcolato.

    griglia_quantistica = list(ParameterGrid(iperparametri_quantistici))
    griglia_classica = list(ParameterGrid(iperparametri_classici))

    # DICHIARAZIONE DELLE VARIABILI DA AGGIORNARE DURANTE L'ADDESTRAMENTO.

    miglior_modello = None 
    miglior_score = -1.0 
    migliori_matrici_gram = None
    miglior_set_adattato = None  
    migliori_iperparametri_quantistici = {} 
    migliori_iperparametri_classici = {} 
    storico_risultati_addestramento = [] # Variabile che serve per esportare i risultati di ogni combinaizone di iperparametri.

    # CREAZIONE DELLA BARRA DI AVANZAMENTO VISIBILE NEL TERMINALE.

    totale_combinazioni= len(griglia_quantistica) * len(griglia_classica)
    with tqdm(total = totale_combinazioni, desc = "Addestramento dei modelli e validazione sugli iperparametri", bar_format = "{l_bar}{bar}| {n_fmt}/{total_fmt}") as barra:
        
        # CICLO PER SCORRERE SUGLI IPERPARAMETRI QUANTISTICI. 

        for combinazione_iperparametri_quantistici in griglia_quantistica:
            
            tempo_inizio_kernel = time.time()

            set_adattato, matrici_gram = modulo_encoding.kernel(set_preparato, numero_features, **combinazione_iperparametri_quantistici)
            matrice_gram_train, matrice_gram_val, _ = matrici_gram

            tempo_fine_kernel = time.time()
            tempo_kernel = tempo_fine_kernel - tempo_inizio_kernel # Misuriamo quanto tempo è stato necessario per calcolare il kernel. 

            # CICLO PER SCORRERE SUGLI IPERPARAMETRI CLASSICI.

            for combinazione_iperparametri_classici in griglia_classica:
                
                tempo_inizio_SVM = time.time()

                modello = SVC(kernel = "precomputed", **combinazione_iperparametri_classici) 
                modello.fit(matrice_gram_train, y_train_preparato)
                
                predizioni_val = modello.predict(matrice_gram_val)
                score_val = f1_score(y_val_preparato, predizioni_val, average = "macro")
                
                tempo_fine_svm = time.time()
                tempo_SVM = tempo_fine_svm - tempo_inizio_SVM # Misuriamo quanto tempo è stato necessario per addestrare e validare la SVM.
                tempo_totale_combinazione = tempo_kernel + tempo_SVM # Calcoliamo il tempo totale che è stato necessario per ottenere questo modello.

                # SALVATAGGIO DEI DATI.

                risultato_corrente = {
                    "encoding": nome_encoding,
                    "iperparametri_quantistici": combinazione_iperparametri_quantistici,
                    "iperparametri_classici": combinazione_iperparametri_classici,
                    "F1_score_macro(%)": round(score_val * 100, 2),
                    "tempo_calcolo_kernel(s)": round(tempo_kernel),
                    "tempo_addestramento_validazione_SVM(s)": round(tempo_SVM),
                    "tempo_esecuzione(s)": round(tempo_totale_combinazione)
                }

                storico_risultati_addestramento.append(risultato_corrente)
                
                # AGGIORNAMENTO DEL MODELLO.

                if score_val > miglior_score:
                    miglior_score = score_val
                    miglior_modello = modello
                    migliori_matrici_gram = matrici_gram
                    miglior_set_adattato = set_adattato  
                    migliori_iperparametri_quantistici = combinazione_iperparametri_quantistici 
                    migliori_iperparametri_classici = combinazione_iperparametri_classici

                barra.update(1)

    # STAMPA NEL TERMINALE DELLA MIGLIOR CONFIGURAZIONE DI IPERPARAMETRI TROVATA, E RELATIVO F1-SCORE.

    if migliori_iperparametri_quantistici: # Scrivendo così, diciamo che, in caso non ce ne siano, non viene stampato nulla. 
        for nome_iperparametro, valore_iperparametro in migliori_iperparametri_quantistici.items():
            print(f"Miglior iperparametro quantistico per {nome_iperparametro}: {valore_iperparametro}")
            
    if migliori_iperparametri_classici:
        for nome_iperparametro, valore_iperparametro in migliori_iperparametri_classici.items():
            print(f"Miglior iperparametro classico per {nome_iperparametro}: {valore_iperparametro}")
            
    print(f"F1-Score (Macro) del miglior modello sul validation set: {miglior_score * 100:.2f}%\n")

    # SALVATAGGIO DEL FILE CON I RISULTATI DI OGNI COMBINAZIONE DI IPERPARAMETRI (codice nelle righe fatto interamente da Gemini).
    
    # Salviamo direttamente in un file .xlsx, con le colonne belle larghe per leggere il testo, ed evidenziando la riga del modello vincente.

    # ------------
    df = pd.DataFrame(storico_risultati_addestramento)
    df["iperparametri_quantistici"] = df["iperparametri_quantistici"].astype(str)
    df["iperparametri_classici"] = df["iperparametri_classici"].astype(str)
    nome_file_excel = os.path.join(nome_cartella, "storico_risultati.xlsx")
    with pd.ExcelWriter(nome_file_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Risultati')
        foglio = writer.sheets['Risultati']
        from openpyxl.utils import get_column_letter 
        from openpyxl.styles import PatternFill # <--- Nuova importazione per i colori
        for indice_colonna, nome_colonna in enumerate(df.columns):
            lunghezza_massima = max(
                df[nome_colonna].astype(str).map(len).max(),
                len(nome_colonna)
            ) + 2
            lettera_colonna = get_column_letter(indice_colonna + 1)
            foglio.column_dimensions[lettera_colonna].width = lunghezza_massima
        indice_migliore = df["F1_score_macro(%)"].idxmax()
        riga_excel = indice_migliore + 2 # +1 per l'intestazione, +1 perché Excel conta partendo da 1
        riempimento_giallo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for col in range(1, len(df.columns) + 1):
            foglio.cell(row=riga_excel, column=col).fill = riempimento_giallo
    # ------------
        
    return miglior_set_adattato, migliori_matrici_gram, miglior_modello, storico_risultati_addestramento